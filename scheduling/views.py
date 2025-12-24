# scheduling/views.py
from django.shortcuts import render
from django.db.models import Q
from django.db.models.deletion import ProtectedError
from django.urls import reverse
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.shortcuts import redirect
from datetime import datetime, timedelta
import calendar as month_calendar
from django.utils.timezone import localtime, now
import time
import logging

from users.models import UserProfile
from .models import Shift, SchedulingWindow, WorkAvailability, Store
from django.utils.dateparse import parse_date

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

from django.views.decorators.csrf import csrf_exempt

import json
import math

logger = logging.getLogger(__name__)


def pick_text_color(hex_color):
    value = hex_color.lstrip("#")
    if len(value) != 6:
        return "#0b3a6b"
    try:
        r = int(value[0:2], 16)
        g = int(value[2:4], 16)
        b = int(value[4:6], 16)
    except ValueError:
        return "#0b3a6b"
    luminance = (0.299 * r + 0.587 * g + 0.114 * b)
    return "#0b3a6b" if luminance > 160 else "#ffffff"


def get_store_display(shift):
    if shift.store_id and shift.store:
        color = shift.store.color
        return shift.store.name, color, pick_text_color(color)
    return "", "#e5e7eb", "#374151"

def get_national_holidays(year):
    lunar_based = {
        2024: {
            "2024-02-10": "農曆新年",
            "2024-06-10": "端午節",
        },
        2025: {
            "2025-01-29": "農曆新年",
            "2025-05-31": "端午節",
        },
        2026: {
            "2026-02-15": "小年夜",
            "2026-02-16": "除夕",
            "2026-02-17": "初一",
            "2026-02-18": "初二",
            "2026-02-19": "初三",
            "2026-02-20": "初四",
            "2026-02-21": "初五",
            "2026-06-19": "端午節",
        },
    }
    holidays = {
        f"{year}-01-01": "元旦",
        f"{year}-02-28": "和平紀念日",
        f"{year}-04-04": "兒童節",
        f"{year}-04-05": "清明節",
        f"{year}-05-01": "勞動節",
        f"{year}-10-10": "國慶日",
        f"{year}-12-25": "行憲紀念日",
    }
    for date_str, name in lunar_based.get(year, {}).items():
        holidays[date_str] = name
    return holidays

def build_holiday_map(dates):
    years = {d.year for d in dates}
    holidays = {}
    for year in years:
        holidays.update(get_national_holidays(year))
    return holidays

def is_manager(user):
    try:
        return user.is_authenticated and user.userprofile.is_manager()
    except UserProfile.DoesNotExist:
        return False


def is_worker(user):
    try:
        return user.is_authenticated and user.userprofile.role == "worker"
    except UserProfile.DoesNotExist:
        return False


def get_active_window():
    latest = SchedulingWindow.objects.order_by("-created_at").first()
    if latest:
        return (
            latest.start_date,
            latest.end_date,
            True,
            latest.allow_worker_view,
            latest.allow_worker_edit_shifts,
            latest.allow_worker_register,
            latest.break_rules or [],
        )
    today = localtime(now()).date()
    start = today - timedelta(days=today.weekday())
    end = start + timedelta(days=6)
    return start, end, False, False, False, False, []


BREAK_MINUTE_OPTIONS = {0, 30, 60, 90, 120}


def parse_min_hours(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        hours = float(value)
    else:
        text = str(value).strip()
        if not text:
            return None
        if ":" in text:
            parts = text.split(":")
            if len(parts) != 2:
                return None
            try:
                hours_part = int(parts[0])
                minutes_part = int(parts[1])
            except ValueError:
                return None
            if hours_part < 0 or minutes_part < 0 or minutes_part >= 60:
                return None
            hours = hours_part + (minutes_part / 60)
        else:
            try:
                hours = float(text)
            except ValueError:
                return None
    if hours <= 0 or math.isnan(hours):
        return None
    return hours


def minutes_to_time_str(total_minutes):
    hours = total_minutes // 60
    minutes = total_minutes % 60
    return f"{hours:02d}:{minutes:02d}"


def normalize_break_rules(raw_rules):
    normalized = []
    for rule in raw_rules or []:
        min_hours = parse_min_hours(rule.get("min_hours"))
        if min_hours is None:
            continue
        try:
            break_minutes = int(rule.get("break_minutes"))
        except (TypeError, ValueError):
            continue
        if break_minutes not in BREAK_MINUTE_OPTIONS:
            continue
        normalized.append({
            "min_hours": min_hours,
            "break_minutes": break_minutes,
        })
    return sorted(normalized, key=lambda r: r["min_hours"])


def parse_break_minutes(value):
    if value is None or value == "":
        return None
    try:
        minutes = int(value)
    except (TypeError, ValueError):
        return None
    if minutes not in BREAK_MINUTE_OPTIONS:
        return None
    return minutes


def calculate_break_minutes(break_rules, start_time, end_time):
    start_min = start_time.hour * 60 + start_time.minute
    end_min = end_time.hour * 60 + end_time.minute
    if end_min <= start_min:
        return 0
    duration = end_min - start_min
    applied = 0
    for rule in normalize_break_rules(break_rules):
        if duration > int(rule["min_hours"] * 60):
            applied = max(applied, rule["break_minutes"])
    return applied


@login_required
@user_passes_test(is_manager)
def scheduling_list(request):
    return redirect("scheduling:timeline")


@login_required
def scheduling_timeline(request):
    try:
        profile = request.user.userprofile
    except UserProfile.DoesNotExist:
        return redirect("users:login")

    is_manager_user = profile.is_manager()
    is_worker_user = profile.role == "worker"
    if not is_manager_user and not is_worker_user:
        return redirect("users:login")

    _, _, _, allow_worker_view, allow_worker_edit_shifts, _, break_rules = get_active_window()
    if not is_manager_user and not allow_worker_view:
        return render(
            request,
            "scheduling/timeline_locked.html",
            {"allow_worker_view": allow_worker_view},
        )

    read_only = not is_manager_user
    hide_store_filter = is_worker_user
    hide_store_info = False
    worker_edit_closed = is_worker_user and not allow_worker_edit_shifts
    t0 = time.perf_counter()
    view = request.GET.get("view", "week")
    if view not in ("day", "week", "month"):
        view = "week"

    date_str = request.GET.get("date")
    store_ids = request.GET.getlist("store")
    show_empty_rows = request.GET.get("show_empty") == "1"
    selected_store_ids = []
    selected_unassigned = False
    store_query = None
    if store_ids:
        unassigned = "unassigned" in store_ids
        selected_unassigned = unassigned
        numeric_ids = []
        for value in store_ids:
            if value.isdigit():
                numeric_ids.append(int(value))
        selected_store_ids = numeric_ids
        if numeric_ids and unassigned:
            store_query = Q(store_id__in=numeric_ids) | Q(store__isnull=True)
        elif numeric_ids:
            store_query = Q(store_id__in=numeric_ids)
        elif unassigned:
            store_query = Q(store__isnull=True)
    date = parse_date(date_str) if date_str else None
    if not date:
        date = localtime(now()).date()

    month_str = request.GET.get("month")
    month_date = None
    if month_str:
        try:
            year, month = month_str.split("-")
            month_date = datetime(int(year), int(month), 1).date()
        except ValueError:
            month_date = None
    if not month_date:
        month_date = date.replace(day=1)
    if view == "month" and not date_str:
        date = month_date

    if not is_manager_user:
        show_empty_rows = False

    hide_empty_rows = not show_empty_rows

    def build_display_name(profile):
        return profile.display_name()

    stores = Store.objects.all()
    store_ids = request.GET.getlist("store")
    store_query = None
    selected_store_ids = []
    selected_unassigned = False
    if store_ids:
        unassigned = "unassigned" in store_ids
        selected_unassigned = unassigned
        numeric_ids = []
        for value in store_ids:
            if value.isdigit():
                numeric_ids.append(int(value))
        selected_store_ids = numeric_ids
        if numeric_ids and unassigned:
            store_query = Q(store_id__in=numeric_ids) | Q(store__isnull=True)
        elif numeric_ids:
            store_query = Q(store_id__in=numeric_ids)
        elif unassigned:
            store_query = Q(store__isnull=True)
    shift_create_url = reverse("scheduling:shift_create")
    shift_update_url = reverse("scheduling:shift_update")
    shift_delete_url = reverse("scheduling:shift_delete")

    today_str = localtime(now()).date().strftime("%Y-%m-%d")
    def format_minutes(total_minutes):
        hours = total_minutes // 60
        minutes = total_minutes % 60
        return f"{hours:02d}:{minutes:02d}"

    if view == "month":
        _, days_in_month = month_calendar.monthrange(month_date.year, month_date.month)
        day_list = [month_date.replace(day=i) for i in range(1, days_in_month + 1)]
        holiday_map = build_holiday_map(day_list)

        workers = UserProfile.objects.filter(role="worker").select_related("user").order_by(
            "sort_order", "name", "user__username"
        )

        shifts_qs = Shift.objects.filter(date__in=day_list, is_published=True)
        if store_query is not None:
            shifts_qs = shifts_qs.filter(store_query)
        shifts = (
            shifts_qs.select_related("employee__user", "store")
            .order_by("start_time")
        )

        scheduled_minutes_by_employee = {}
        for s in shifts:
            if not s.store_id:
                continue
            start_min = s.start_time.hour * 60 + s.start_time.minute
            end_min = s.end_time.hour * 60 + s.end_time.minute
            if end_min <= start_min:
                end_min += 24 * 60
            scheduled_minutes_by_employee[s.employee_id] = (
                scheduled_minutes_by_employee.get(s.employee_id, 0) + (end_min - start_min)
            )

        by_employee_date = {}
        for s in shifts:
            by_employee_date.setdefault(s.employee_id, {}).setdefault(s.date, []).append(s)

        month_rows = []
        for worker in workers:
            day_cells = []
            for d in day_list:
                items = []
                for s in by_employee_date.get(worker.id, {}).get(d, []):
                    store_name, store_color, store_text_color = get_store_display(s)
                    items.append({
                        "id": s.id,
                        "date": d.strftime("%Y-%m-%d"),
                        "start": s.start_time.strftime("%H:%M"),
                        "end": s.end_time.strftime("%H:%M"),
                        "note": s.note,
                        "break_minutes": s.break_minutes,
                        "store_id": s.store_id,
                        "store_name": store_name,
                        "store_color": store_color,
                        "store_text_color": store_text_color,
                        "label": f"{s.start_time.strftime('%H:%M')}-{s.end_time.strftime('%H:%M')} {store_name}".strip(),
                        "label_no_store": f"{s.start_time.strftime('%H:%M')}-{s.end_time.strftime('%H:%M')}",
                    })
                date_str = d.strftime("%Y-%m-%d")
                day_cells.append({
                    "date": date_str,
                    "is_weekend": d.weekday() >= 5,
                    "holiday_name": holiday_map.get(date_str),
                    "shifts": items,
                })
            if hide_empty_rows and not any(cell["shifts"] for cell in day_cells):
                continue
            month_rows.append({
                "employee": worker,
                "display_name": build_display_name(worker),
                "scheduled_hours": format_minutes(scheduled_minutes_by_employee.get(worker.id, 0)),
                "day_cells": day_cells,
            })

        weekday_labels = ["一", "二", "三", "四", "五", "六", "日"]
        response = render(request, "scheduling/timeline.html", {
            "date": date,
            "view": view,
            "month_date": month_date,
            "month_days": [{
                "day": d.day,
                "date": d.strftime("%Y-%m-%d"),
                "is_weekend": d.weekday() >= 5,
                "holiday_name": holiday_map.get(d.strftime("%Y-%m-%d")),
                "weekday_label": weekday_labels[d.weekday()],
            } for d in day_list],
            "month_rows": month_rows,
            "month_value": month_date.strftime("%Y-%m"),
            "read_only": read_only,
            "shift_create_url": shift_create_url,
            "shift_update_url": shift_update_url,
            "shift_delete_url": shift_delete_url,
            "today_str": today_str,
            "stores": stores,
            "selected_store_ids": selected_store_ids,
            "selected_unassigned": selected_unassigned,
            "hide_store_filter": hide_store_filter,
            "hide_store_info": hide_store_info,
            "worker_edit_closed": worker_edit_closed,
            "can_manage_store": is_manager_user,
            "show_empty_rows": show_empty_rows,
            "break_rules_json": json.dumps(normalize_break_rules(break_rules)),
        })
        logger.info("timeline view=%s total_ms=%.1f", view, (time.perf_counter() - t0) * 1000)
        return response

    day_headers = None
    weekday_labels = ["一", "二", "三", "四", "五", "六", "日"]
    if view == "week":
        start_date = date - timedelta(days=date.weekday())
        date_range = [start_date + timedelta(days=i) for i in range(7)]
    else:
        date_range = [date]

    holiday_map = build_holiday_map(date_range)
    if view in ("week", "day"):
        day_headers = [{
            "label": f"{d.strftime('%m/%d')}（{weekday_labels[d.weekday()]}）",
            "date": d,
            "date_str": d.strftime("%Y-%m-%d"),
            "is_weekend": d.weekday() >= 5,
            "holiday_name": holiday_map.get(d.strftime("%Y-%m-%d")),
        } for d in date_range]

    base_start = 8 * 60
    base_end = 24 * 60
    total_minutes = base_end - base_start

    workers = UserProfile.objects.filter(role="worker").select_related("user").order_by(
        "sort_order", "name", "user__username"
    )

    shifts_qs = Shift.objects.filter(date__in=date_range, is_published=True)
    if store_query is not None:
        shifts_qs = shifts_qs.filter(store_query)
    shifts = (
        shifts_qs.select_related("employee__user", "store")
        .order_by("start_time")
    )

    scheduled_minutes_by_employee = {}
    for s in shifts:
        if not s.store_id:
            continue
        start_min = s.start_time.hour * 60 + s.start_time.minute
        end_min = s.end_time.hour * 60 + s.end_time.minute
        if end_min <= start_min:
            end_min += 24 * 60
        scheduled_minutes_by_employee[s.employee_id] = (
            scheduled_minutes_by_employee.get(s.employee_id, 0) + (end_min - start_min)
        )

    by_employee_date = {}
    for s in shifts:
        by_employee_date.setdefault(s.employee_id, {}).setdefault(s.date, []).append(s)

    rows = []
    for worker in workers:
        day_entries = []
        for d in date_range:
            items = []
            for s in by_employee_date.get(worker.id, {}).get(d, []):
                store_name, store_color, store_text_color = get_store_display(s)
                start_min = s.start_time.hour * 60 + s.start_time.minute
                end_min = s.end_time.hour * 60 + s.end_time.minute
                if end_min <= start_min:
                    end_min += 24 * 60

                display_start = max(start_min, base_start)
                display_end = min(end_min, base_end)
                if display_end <= base_start or display_start >= base_end:
                    continue

                offset = display_start - base_start
                width = display_end - display_start
                items.append({
                    "id": s.id,
                    "date": d,
                    "start_label": s.start_time.strftime("%H:%M"),
                    "end_label": s.end_time.strftime("%H:%M"),
                    "note": s.note,
                    "break_minutes": s.break_minutes,
                    "store_id": s.store_id,
                    "store_name": store_name,
                    "store_color": store_color,
                    "store_text_color": store_text_color,
                    "offset_pct": round(offset / total_minutes * 100, 4),
                    "width_pct": round(width / total_minutes * 100, 4),
                })

            day_entries.append({
                "date": d,
                "label": d.strftime("%m/%d"),
                "shifts": items,
                "date_str": d.strftime("%Y-%m-%d"),
                "is_weekend": d.weekday() >= 5,
                "holiday_name": holiday_map.get(d.strftime("%Y-%m-%d")),
            })

        if hide_empty_rows and not any(entry["shifts"] for entry in day_entries):
            continue

        rows.append({
            "employee": worker,
            "display_name": build_display_name(worker),
            "scheduled_hours": format_minutes(scheduled_minutes_by_employee.get(worker.id, 0)),
            "days": day_entries,
        })

    hours = list(range(8, 24))

    response = render(request, "scheduling/timeline.html", {
        "date": date,
        "rows": rows,
        "hours": hours,
        "view": view,
        "day_headers": day_headers,
        "month_value": date.strftime("%Y-%m"),
        "read_only": read_only,
        "shift_create_url": shift_create_url,
        "shift_update_url": shift_update_url,
        "shift_delete_url": shift_delete_url,
        "today_str": today_str,
        "stores": stores,
        "selected_store_ids": selected_store_ids,
        "selected_unassigned": selected_unassigned,
        "hide_store_filter": hide_store_filter,
        "hide_store_info": hide_store_info,
        "show_empty_rows": show_empty_rows,
        "worker_edit_closed": worker_edit_closed,
        "can_manage_store": is_manager_user,
        "break_rules_json": json.dumps(normalize_break_rules(break_rules)),
    })
    logger.info("timeline view=%s total_ms=%.1f", view, (time.perf_counter() - t0) * 1000)
    return response



@login_required
@user_passes_test(is_manager)
def export_excel_view(request):
    """
    匯出目前所有 Shift 為 Excel。
    """
    shifts = Shift.objects.select_related('employee__user', 'store').all().order_by('date', 'start_time')

    wb = Workbook()
    ws = wb.active
    ws.title = "班表"

    # 標題列
    headers = ["日期", "員工姓名", "店別", "開始時間", "結束時間", "備註", "是否發佈"]
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center")

    for col, title in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.font = header_font
        cell.alignment = center
        cell.fill = header_fill

    # 資料列
    row_num = 2
    for s in shifts:
        start_dt = datetime.combine(s.date, s.start_time)
        end_dt = datetime.combine(s.date, s.end_time)
        if end_dt < start_dt:
            end_dt += timedelta(days=1)

        ws.cell(row=row_num, column=1, value=s.date.strftime("%Y-%m-%d"))
        ws.cell(row=row_num, column=2, value=s.employee.display_name())
        ws.cell(row=row_num, column=3, value=s.store.name if s.store else "")
        ws.cell(row=row_num, column=4, value=s.start_time.strftime("%H:%M"))
        ws.cell(row=row_num, column=5, value=s.end_time.strftime("%H:%M"))
        ws.cell(row=row_num, column=6, value=s.note)
        ws.cell(row=row_num, column=7, value="是" if s.is_published else "否")

        row_num += 1

    # 欄寬自動調整
    for column_cells in ws.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter
        for cell in column_cells:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column_letter].width = max_length + 2

    # 回應
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename_time = localtime(now()).strftime("%Y%m%d_%H%M%S")
    response['Content-Disposition'] = f'attachment; filename="shift_report_{filename_time}.xlsx"'

    wb.save(response)
    return response

@csrf_exempt
def create_shift(request):
    if not request.user.is_authenticated or not is_manager(request.user):
        return JsonResponse({"ok": False, "error": "僅限店長操作"}, status=403)
    data = json.loads(request.body)
    employee_id = data.get("employee_id")
    store_id = data.get("store_id") or None
    date_str = data.get("date")
    start_str = data.get("start")
    end_str = data.get("end")
    raw_break_minutes = data.get("break_minutes")
    note = (data.get("note") or "").strip()

    if not employee_id or not date_str or not start_str or not end_str:
        return JsonResponse({"ok": False, "error": "missing fields"}, status=400)

    date = parse_date(date_str)
    if not date:
        return JsonResponse({"ok": False, "error": "invalid date"}, status=400)

    try:
        start_time = datetime.strptime(start_str, "%H:%M").time()
        end_time = datetime.strptime(end_str, "%H:%M").time()
    except ValueError:
        return JsonResponse({"ok": False, "error": "invalid time"}, status=400)

    start_min = start_time.hour * 60 + start_time.minute
    end_min = end_time.hour * 60 + end_time.minute
    if end_min <= start_min:
        return JsonResponse({"ok": False, "error": "invalid range"}, status=400)

    if store_id and not Store.objects.filter(id=store_id).exists():
        return JsonResponse({"ok": False, "error": "invalid store"}, status=400)

    break_minutes = parse_break_minutes(raw_break_minutes)
    if raw_break_minutes not in (None, "") and break_minutes is None:
        return JsonResponse({"ok": False, "error": "invalid break minutes"}, status=400)
    if break_minutes is None:
        _, _, _, _, _, _, break_rules = get_active_window()
        break_minutes = calculate_break_minutes(break_rules, start_time, end_time)

    conflict = Shift.objects.filter(
        employee_id=employee_id,
        date=date,
        start_time__lt=end_time,
        end_time__gt=start_time,
    ).exists()
    if conflict:
        return JsonResponse(
            {"ok": False, "error": "排班時間重疊，請重新確認。"},
            status=400,
        )

    shift = Shift.objects.create(
        employee_id=employee_id,
        store_id=store_id,
        date=date,
        start_time=start_time,
        end_time=end_time,
        break_minutes=break_minutes,
        is_published=True,
        note=note,
    )

    return JsonResponse({"ok": True, "id": shift.id})

@csrf_exempt
def delete_shift(request):
    if not request.user.is_authenticated or not is_manager(request.user):
        return JsonResponse({"ok": False, "error": "僅限店長操作"}, status=403)
    data = json.loads(request.body)
    shift_id = data.get("id")

    Shift.objects.filter(id=shift_id).delete()
    return JsonResponse({"ok": True})

@csrf_exempt
def update_shift(request):
    if not request.user.is_authenticated or not is_manager(request.user):
        return JsonResponse({"ok": False, "error": "僅限店長操作"}, status=403)
    data = json.loads(request.body)

    shift = Shift.objects.get(id=data["id"])
    store_id = data.get("store_id")
    split_start_str = data.get("split_start")
    split_end_str = data.get("split_end")
    split_store_id = data.get("split_store_id")
    note = data.get("note")
    raw_break_minutes = data.get("break_minutes")
    try:
        start_time = datetime.strptime(data["start"], "%H:%M").time()
        end_time = datetime.strptime(data["end"], "%H:%M").time()
    except ValueError:
        return JsonResponse({"ok": False, "error": "invalid time"}, status=400)

    if (end_time.hour * 60 + end_time.minute) <= (start_time.hour * 60 + start_time.minute):
        return JsonResponse({"ok": False, "error": "invalid range"}, status=400)

    if store_id in ("", "null", "none"):
        store_id = None
    if store_id and not Store.objects.filter(id=store_id).exists():
        return JsonResponse({"ok": False, "error": "invalid store"}, status=400)

    break_minutes = parse_break_minutes(raw_break_minutes)
    if raw_break_minutes not in (None, "") and break_minutes is None:
        return JsonResponse({"ok": False, "error": "invalid break minutes"}, status=400)

    split_start_time = None
    split_end_time = None
    if split_start_str or split_end_str or split_store_id:
        if not split_start_str or not split_end_str or not split_store_id:
            return JsonResponse({"ok": False, "error": "invalid split fields"}, status=400)
        try:
            split_start_time = datetime.strptime(split_start_str, "%H:%M").time()
            split_end_time = datetime.strptime(split_end_str, "%H:%M").time()
        except ValueError:
            return JsonResponse({"ok": False, "error": "invalid split time"}, status=400)
        if (split_end_time.hour * 60 + split_end_time.minute) <= (split_start_time.hour * 60 + split_start_time.minute):
            return JsonResponse({"ok": False, "error": "invalid split range"}, status=400)
        if not Store.objects.filter(id=split_store_id).exists():
            return JsonResponse({"ok": False, "error": "invalid split store"}, status=400)
        if split_start_time < end_time and split_end_time > start_time:
            return JsonResponse(
                {"ok": False, "error": "排班時間重疊，請重新確認。"},
                status=400,
            )

    conflict = Shift.objects.filter(
        employee_id=shift.employee_id,
        date=shift.date,
        start_time__lt=end_time,
        end_time__gt=start_time,
    ).exclude(id=shift.id).exists()
    if conflict:
        return JsonResponse(
            {"ok": False, "error": "排班時間重疊，請重新確認。"},
            status=400,
        )

    if split_start_time and split_end_time:
        split_conflict = Shift.objects.filter(
            employee_id=shift.employee_id,
            date=shift.date,
            start_time__lt=split_end_time,
            end_time__gt=split_start_time,
        ).exclude(id=shift.id).exists()
        if split_conflict:
            return JsonResponse(
                {"ok": False, "error": "排班時間重疊，請重新確認。"},
                status=400,
            )

    shift.start_time = start_time
    shift.end_time = end_time
    shift.store_id = store_id
    if break_minutes is not None:
        shift.break_minutes = break_minutes
    if note is not None:
        shift.note = note.strip()
    shift.save()

    if split_start_time and split_end_time:
        Shift.objects.create(
            employee_id=shift.employee_id,
            store_id=split_store_id,
            date=shift.date,
            start_time=split_start_time,
            end_time=split_end_time,
            is_published=True,
            break_minutes=0,
        )

    return JsonResponse({"ok": True})


@login_required
@user_passes_test(is_manager)
def manage_window(request):
    start_date, end_date, has_manager_window, allow_worker_view, allow_worker_edit_shifts, allow_worker_register, break_rules = get_active_window()
    stores = Store.objects.all()
    current_break_rules = normalize_break_rules(break_rules)
    break_rules_form = []
    for rule in current_break_rules:
        minutes = int(rule["min_hours"] * 60)
        break_rules_form.append({
            "min_hours": rule["min_hours"],
            "min_time": minutes_to_time_str(minutes),
            "break_minutes": rule["break_minutes"],
        })
    break_threshold_options = [minutes_to_time_str(m) for m in range(30, (12 * 60) + 1, 30)]

    if request.method == "POST":
        store_name = request.POST.get("store_name")
        delete_store_id = request.POST.get("delete_store_id")
        store_id = request.POST.get("store_id")
        store_color = request.POST.get("store_color")
        if delete_store_id:
            try:
                deleted, _ = Store.objects.filter(id=delete_store_id).delete()
                if deleted:
                    messages.success(request, "店別已刪除。")
                else:
                    messages.error(request, "找不到店別。")
            except ProtectedError:
                messages.error(request, "店別已有班表，無法刪除。")
            return redirect("scheduling:manage_window")
        if store_id and store_color:
            updated = Store.objects.filter(id=store_id).update(color=store_color)
            if updated:
                messages.success(request, "店別顏色已更新。")
            else:
                messages.error(request, "找不到店別。")
            return redirect("scheduling:manage_window")
        if store_name is not None:
            store_name = store_name.strip()
            if not store_name:
                messages.error(request, "請輸入店別名稱。")
            elif Store.objects.filter(name=store_name).exists():
                messages.error(request, "店別已存在。")
            else:
                color_value = request.POST.get("store_color", "#cfe8ff")
                Store.objects.create(name=store_name, color=color_value)
                messages.success(request, "店別已新增。")
                return redirect("scheduling:manage_window")
        action = request.POST.get("action")
        if action == "update_dates":
            start_str = request.POST.get("start_date")
            end_str = request.POST.get("end_date")
            allow_worker_view = allow_worker_view
            allow_worker_edit_shifts = allow_worker_edit_shifts
            allow_worker_register = allow_worker_register
            start = parse_date(start_str) if start_str else None
            end = parse_date(end_str) if end_str else None
            if not start or not end:
                messages.error(request, "請填寫完整日期。")
            elif end < start:
                messages.error(request, "開始日期不可超過結束日期。")
            else:
                SchedulingWindow.objects.create(
                    start_date=start,
                    end_date=end,
                    allow_worker_view=allow_worker_view,
                    allow_worker_edit_shifts=allow_worker_edit_shifts,
                    allow_worker_register=allow_worker_register,
                    break_rules=current_break_rules,
                )
                messages.success(request, "可排班日期設定已更新。")
                return redirect("scheduling:manage_window")
            start_date, end_date = start or start_date, end or end_date
        elif action == "update_permissions":
            allow_worker_view = request.POST.get("allow_worker_view") == "on"
            allow_worker_edit_shifts = request.POST.get("allow_worker_edit_shifts") == "on"
            allow_worker_register = request.POST.get("allow_worker_register") == "on"
            SchedulingWindow.objects.create(
                start_date=start_date,
                end_date=end_date,
                allow_worker_view=allow_worker_view,
                allow_worker_edit_shifts=allow_worker_edit_shifts,
                allow_worker_register=allow_worker_register,
                break_rules=current_break_rules,
            )
            messages.success(request, "員工設定已更新。")
            return redirect("scheduling:manage_window")
        elif action == "update_break_rules":
            thresholds = request.POST.getlist("break_threshold_hours")
            minutes_list = request.POST.getlist("break_minutes")
            raw_rules = []
            has_partial = False
            for idx in range(min(len(thresholds), len(minutes_list))):
                threshold = (thresholds[idx] or "").strip()
                minutes = (minutes_list[idx] or "").strip()
                if not threshold:
                    if minutes and minutes != "0":
                        has_partial = True
                    continue
                if not minutes:
                    has_partial = True
                    continue
                raw_rules.append({
                    "min_hours": threshold,
                    "break_minutes": minutes,
                })
            new_rules = normalize_break_rules(raw_rules)
            if has_partial or len(raw_rules) != len(new_rules):
                messages.error(request, "排班休息時間設定有誤，請確認數值。")
                return redirect("scheduling:manage_window")
            SchedulingWindow.objects.create(
                start_date=start_date,
                end_date=end_date,
                allow_worker_view=allow_worker_view,
                allow_worker_edit_shifts=allow_worker_edit_shifts,
                allow_worker_register=allow_worker_register,
                break_rules=new_rules,
            )
            messages.success(request, "排班休息時間已更新。")
            return redirect("scheduling:manage_window")

    return render(
        request,
        "scheduling/manage_window.html",
        {
            "start_date": start_date,
            "end_date": end_date,
            "has_manager_window": has_manager_window,
            "allow_worker_view": allow_worker_view,
            "allow_worker_edit_shifts": allow_worker_edit_shifts,
            "allow_worker_register": allow_worker_register,
            "stores": stores,
            "break_rules": break_rules_form,
            "break_threshold_options": break_threshold_options,
        },
    )


@login_required
def worker_schedule(request):
    try:
        profile = request.user.userprofile
    except UserProfile.DoesNotExist:
        return redirect("users:login")

    if profile.is_manager():
        return redirect("scheduling:timeline")

    start_date, end_date, has_manager_window, allow_worker_view, allow_worker_edit_shifts, _, break_rules = get_active_window()
    view = request.GET.get("view", "week")
    if view not in ("day", "week", "month"):
        view = "week"
    date_str = request.GET.get("date")
    date = parse_date(date_str) if date_str else None
    if not date:
        date = localtime(now()).date()

    month_str = request.GET.get("month")
    month_date = None
    if month_str:
        try:
            year, month = month_str.split("-")
            month_date = datetime(int(year), int(month), 1).date()
        except ValueError:
            month_date = None
    if not month_date:
        month_date = date.replace(day=1)
    if view == "month" and not date_str:
        date = month_date

    weekday_labels = ["一", "二", "三", "四", "五", "六", "日"]
    day_headers = None
    month_days = None
    if view == "week":
        week_start = date - timedelta(days=date.weekday())
        date_range = [week_start + timedelta(days=i) for i in range(7)]
    elif view == "day":
        date_range = [date]
    else:
        _, days_in_month = month_calendar.monthrange(month_date.year, month_date.month)
        date_range = [month_date.replace(day=i) for i in range(1, days_in_month + 1)]

    holiday_map = build_holiday_map(date_range)
    if view in ("week", "day"):
        day_headers = [{
            "label": f"{d.strftime('%m/%d')}（{weekday_labels[d.weekday()]}）",
            "date": d,
            "date_str": d.strftime("%Y-%m-%d"),
            "is_weekend": d.weekday() >= 5,
            "holiday_name": holiday_map.get(d.strftime("%Y-%m-%d")),
        } for d in date_range]
    else:
        month_days = [{
            "day": d.day,
            "date": d.strftime("%Y-%m-%d"),
            "is_weekend": d.weekday() >= 5,
            "holiday_name": holiday_map.get(d.strftime("%Y-%m-%d")),
            "weekday_label": weekday_labels[d.weekday()],
        } for d in date_range]

    if allow_worker_view:
        workers = UserProfile.objects.filter(role="worker").select_related("user").order_by(
            "sort_order", "name", "user__username"
        )
        shifts = (
            Shift.objects.filter(date__in=date_range, is_published=True)
            .select_related("employee__user", "store")
            .order_by("start_time")
        )
    else:
        workers = [profile]
        shifts = (
            Shift.objects.filter(employee=profile, date__in=date_range, is_published=True)
            .select_related("employee__user", "store")
            .order_by("start_time")
        )

    by_employee_date = {}
    for s in shifts:
        by_employee_date.setdefault(s.employee_id, {}).setdefault(s.date, []).append(s)

    scheduled_minutes = 0
    for s in shifts:
        if s.employee_id != profile.id or not s.store_id:
            continue
        start_min = s.start_time.hour * 60 + s.start_time.minute
        end_min = s.end_time.hour * 60 + s.end_time.minute
        if end_min <= start_min:
            end_min += 24 * 60
        scheduled_minutes += end_min - start_min

    def format_minutes(total_minutes):
        hours = total_minutes // 60
        minutes = total_minutes % 60
        return f"{hours:02d}:{minutes:02d}"

    rows = []
    month_rows = None
    if view == "month":
        month_rows = []
        for worker in workers:
            day_cells = []
            for d in date_range:
                items = []
                for s in by_employee_date.get(worker.id, {}).get(d, []):
                    store_name, store_color, store_text_color = get_store_display(s)
                    items.append({
                        "id": s.id,
                        "date": d.strftime("%Y-%m-%d"),
                        "start": s.start_time.strftime("%H:%M"),
                        "end": s.end_time.strftime("%H:%M"),
                        "note": s.note,
                        "break_minutes": s.break_minutes,
                        "store_id": s.store_id,
                        "store_name": store_name,
                        "store_color": store_color,
                        "store_text_color": store_text_color,
                        "label": f"{s.start_time.strftime('%H:%M')}-{s.end_time.strftime('%H:%M')} {store_name}".strip(),
                        "label_no_store": f"{s.start_time.strftime('%H:%M')}-{s.end_time.strftime('%H:%M')}",
                    })
                date_str = d.strftime("%Y-%m-%d")
                day_cells.append({
                    "date": date_str,
                    "is_weekend": d.weekday() >= 5,
                    "holiday_name": holiday_map.get(date_str),
                    "shifts": items,
                })
            if not any(cell["shifts"] for cell in day_cells) and worker.id != profile.id:
                continue
            month_rows.append({
                "employee": worker,
                "display_name": worker.display_name(),
                "day_cells": day_cells,
            })
    else:
        for worker in workers:
            day_entries = []
            for d in date_range:
                items = []
                for s in by_employee_date.get(worker.id, {}).get(d, []):
                    store_name, store_color, store_text_color = get_store_display(s)
                    items.append({
                        "id": s.id,
                        "date": d,
                        "start_label": s.start_time.strftime("%H:%M"),
                        "end_label": s.end_time.strftime("%H:%M"),
                        "note": s.note,
                        "break_minutes": s.break_minutes,
                        "store_id": s.store_id,
                        "store_name": store_name,
                        "store_color": store_color,
                        "store_text_color": store_text_color,
                    })

                day_entries.append({
                    "date": d,
                    "label": d.strftime("%m/%d"),
                    "shifts": items,
                    "date_str": d.strftime("%Y-%m-%d"),
                    "is_weekend": d.weekday() >= 5,
                    "holiday_name": holiday_map.get(d.strftime("%Y-%m-%d")),
                })

            if not any(entry["shifts"] for entry in day_entries) and worker.id != profile.id:
                continue

            rows.append({
                "employee": worker,
                "display_name": worker.display_name(),
                "days": day_entries,
            })

    shift_create_url = reverse("scheduling:worker_shift_create")
    shift_update_url = reverse("scheduling:worker_shift_update")
    shift_delete_url = reverse("scheduling:worker_shift_delete")
    show_profile_warning = profile.missing_required_info()

    return render(
        request,
        "scheduling/timeline.html",
        {
            "date": date,
            "rows": rows,
            "view": view,
            "day_headers": day_headers,
            "month_value": month_date.strftime("%Y-%m"),
            "month_days": month_days,
            "month_rows": month_rows,
            "read_only": not allow_worker_edit_shifts,
            "hide_view_toggle": False,
            "page_title": "我的班表",
            "allowed_employee_id": profile.id,
            "can_edit_own_only": True,
            "schedule_window_start": start_date,
            "schedule_window_end": end_date,
            "shift_create_url": shift_create_url,
            "shift_update_url": shift_update_url,
            "shift_delete_url": shift_delete_url,
            "today_str": localtime(now()).date().strftime("%Y-%m-%d"),
            "hide_store_filter": True,
            "hide_store_info": False,
            "worker_edit_closed": not allow_worker_edit_shifts,
            "can_manage_store": False,
            "show_profile_warning": show_profile_warning,
            "self_scheduled_hours": format_minutes(scheduled_minutes),
            "break_rules_json": json.dumps(normalize_break_rules(break_rules)),
        },
    )


def worker_shift_edit_allowed():
    _, _, _, _, allow_worker_edit_shifts, _, _ = get_active_window()
    return allow_worker_edit_shifts


@csrf_exempt
@login_required
def create_availability(request):
    try:
        profile = request.user.userprofile
    except UserProfile.DoesNotExist:
        return JsonResponse({"ok": False, "error": "查無使用者資料"}, status=400)
    if not is_worker(request.user):
        return JsonResponse({"ok": False, "error": "僅限員工操作"}, status=403)

    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({"ok": False, "error": "資料格式錯誤"}, status=400)

    date = parse_date(data.get("date"))
    start_str = data.get("start")
    end_str = data.get("end")
    if not date or not start_str or not end_str:
        return JsonResponse({"ok": False, "error": "請填寫完整日期與時間"}, status=400)

    start_time = datetime.strptime(start_str, "%H:%M").time()
    end_time = datetime.strptime(end_str, "%H:%M").time()
    if (end_time.hour * 60 + end_time.minute) <= (start_time.hour * 60 + start_time.minute):
        return JsonResponse({"ok": False, "error": "結束時間需晚於開始時間"}, status=400)

    start_date, end_date, _, _, _, _, break_rules = get_active_window()
    if date < start_date or date > end_date:
        return JsonResponse({"ok": False, "error": "不在可排班的日期區間內"}, status=400)

    conflict = WorkAvailability.objects.filter(
        employee=profile,
        date=date,
        start_time__lt=end_time,
        end_time__gt=start_time,
    ).exists()
    if conflict:
        return JsonResponse({"ok": False, "error": "時段與已填寫的重疊，請重新確認。"}, status=400)

    shift_conflict = Shift.objects.filter(
        employee=profile,
        date=date,
        start_time__lt=end_time,
        end_time__gt=start_time,
        is_published=True,
    ).exists()
    if shift_conflict:
        return JsonResponse({"ok": False, "error": "與已排班時段重疊，請重新確認。"}, status=400)

    avail = WorkAvailability.objects.create(
        employee=profile,
        date=date,
        start_time=start_time,
        end_time=end_time,
    )

    return JsonResponse(
        {
            "ok": True,
            "id": avail.id,
            "date": date.strftime("%Y-%m-%d"),
            "start": start_str,
            "end": end_str,
        }
    )


@csrf_exempt
@login_required
def delete_availability(request):
    try:
        profile = request.user.userprofile
    except UserProfile.DoesNotExist:
        return JsonResponse({"ok": False, "error": "查無使用者資料"}, status=400)
    if not is_worker(request.user):
        return JsonResponse({"ok": False, "error": "僅限員工操作"}, status=403)

    try:
        data = json.loads(request.body)
        avail_id = int(data.get("id"))
    except Exception:
        return JsonResponse({"ok": False, "error": "資料格式錯誤"}, status=400)

    deleted, _ = WorkAvailability.objects.filter(id=avail_id, employee=profile).delete()
    if not deleted:
        return JsonResponse({"ok": False, "error": "找不到資料"}, status=404)

    return JsonResponse({"ok": True})


@csrf_exempt
@login_required
def update_availability(request):
    try:
        profile = request.user.userprofile
    except UserProfile.DoesNotExist:
        return JsonResponse({"ok": False, "error": "查無使用者資料"}, status=400)
    if not is_worker(request.user):
        return JsonResponse({"ok": False, "error": "僅限員工操作"}, status=403)

    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({"ok": False, "error": "資料格式錯誤"}, status=400)

    avail_id = data.get("id")
    start_str = data.get("start")
    end_str = data.get("end")
    if not avail_id or not start_str or not end_str:
        return JsonResponse({"ok": False, "error": "請填寫完整時間"}, status=400)

    try:
        avail = WorkAvailability.objects.get(id=avail_id, employee=profile)
    except WorkAvailability.DoesNotExist:
        return JsonResponse({"ok": False, "error": "找不到資料"}, status=404)

    try:
        start_time = datetime.strptime(start_str, "%H:%M").time()
        end_time = datetime.strptime(end_str, "%H:%M").time()
    except ValueError:
        return JsonResponse({"ok": False, "error": "時間格式錯誤"}, status=400)

    if (end_time.hour * 60 + end_time.minute) <= (start_time.hour * 60 + start_time.minute):
        return JsonResponse({"ok": False, "error": "結束時間需晚於開始時間"}, status=400)

    start_date, end_date, _, _, _, _, _ = get_active_window()
    if avail.date < start_date or avail.date > end_date:
        return JsonResponse({"ok": False, "error": "不在可排班的日期區間內"}, status=400)

    conflict = WorkAvailability.objects.filter(
        employee=profile,
        date=avail.date,
        start_time__lt=end_time,
        end_time__gt=start_time,
    ).exclude(id=avail.id).exists()
    if conflict:
        return JsonResponse({"ok": False, "error": "時段與已填寫的重疊，請重新確認。"}, status=400)

    shift_conflict = Shift.objects.filter(
        employee=profile,
        date=avail.date,
        start_time__lt=end_time,
        end_time__gt=start_time,
        is_published=True,
    ).exists()
    if shift_conflict:
        return JsonResponse({"ok": False, "error": "與已排班時段重疊，請重新確認。"}, status=400)

    avail.start_time = start_time
    avail.end_time = end_time
    avail.save(update_fields=["start_time", "end_time"])

    return JsonResponse({"ok": True})


@csrf_exempt
@login_required
def create_worker_shift(request):
    try:
        profile = request.user.userprofile
    except UserProfile.DoesNotExist:
        return JsonResponse({"ok": False, "error": "查無使用者資料"}, status=400)
    if not is_worker(request.user):
        return JsonResponse({"ok": False, "error": "僅限員工操作"}, status=403)
    if not worker_shift_edit_allowed():
        return JsonResponse({"ok": False, "error": "目前未開放員工排班"}, status=403)

    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({"ok": False, "error": "資料格式錯誤"}, status=400)

    employee_id = data.get("employee_id")
    if employee_id and str(employee_id) != str(profile.id):
        return JsonResponse({"ok": False, "error": "僅能編輯自己的班表"}, status=403)

    date_str = data.get("date")
    start_str = data.get("start")
    end_str = data.get("end")
    raw_break_minutes = data.get("break_minutes")
    if not date_str or not start_str or not end_str:
        return JsonResponse({"ok": False, "error": "請填寫完整日期與時間"}, status=400)

    date = parse_date(date_str)
    if not date:
        return JsonResponse({"ok": False, "error": "日期格式錯誤"}, status=400)

    try:
        start_time = datetime.strptime(start_str, "%H:%M").time()
        end_time = datetime.strptime(end_str, "%H:%M").time()
    except ValueError:
        return JsonResponse({"ok": False, "error": "時間格式錯誤"}, status=400)

    if (end_time.hour * 60 + end_time.minute) <= (start_time.hour * 60 + start_time.minute):
        return JsonResponse({"ok": False, "error": "結束時間需晚於開始時間"}, status=400)

    start_date, end_date, _, _, _, _, _ = get_active_window()
    if date < start_date or date > end_date:
        return JsonResponse({"ok": False, "error": "不在可排班的日期區間內"}, status=400)

    conflict = Shift.objects.filter(
        employee=profile,
        date=date,
        start_time__lt=end_time,
        end_time__gt=start_time,
        is_published=True,
    ).exists()
    if conflict:
        return JsonResponse({"ok": False, "error": "排班時間重疊，請重新確認。"}, status=400)

    break_minutes = parse_break_minutes(raw_break_minutes)
    if raw_break_minutes not in (None, "") and break_minutes is None:
        return JsonResponse({"ok": False, "error": "休息時間格式錯誤"}, status=400)
    if break_minutes is None:
        break_minutes = calculate_break_minutes(break_rules, start_time, end_time)

    shift = Shift.objects.create(
        employee=profile,
        store_id=None,
        date=date,
        start_time=start_time,
        end_time=end_time,
        break_minutes=break_minutes,
        is_published=True,
    )

    return JsonResponse({"ok": True, "id": shift.id})


@csrf_exempt
@login_required
def update_worker_shift(request):
    try:
        profile = request.user.userprofile
    except UserProfile.DoesNotExist:
        return JsonResponse({"ok": False, "error": "查無使用者資料"}, status=400)
    if not is_worker(request.user):
        return JsonResponse({"ok": False, "error": "僅限員工操作"}, status=403)
    if not worker_shift_edit_allowed():
        return JsonResponse({"ok": False, "error": "目前未開放員工排班"}, status=403)

    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({"ok": False, "error": "資料格式錯誤"}, status=400)

    shift_id = data.get("id")
    start_str = data.get("start")
    end_str = data.get("end")
    raw_break_minutes = data.get("break_minutes")
    if not shift_id or not start_str or not end_str:
        return JsonResponse({"ok": False, "error": "請填寫完整時間"}, status=400)

    try:
        shift = Shift.objects.get(id=shift_id, employee=profile, is_published=True)
    except Shift.DoesNotExist:
        return JsonResponse({"ok": False, "error": "找不到資料"}, status=404)
    if shift.store_id:
        return JsonResponse({"ok": False, "error": "店長排定班表不可修改"}, status=403)

    try:
        start_time = datetime.strptime(start_str, "%H:%M").time()
        end_time = datetime.strptime(end_str, "%H:%M").time()
    except ValueError:
        return JsonResponse({"ok": False, "error": "時間格式錯誤"}, status=400)

    if (end_time.hour * 60 + end_time.minute) <= (start_time.hour * 60 + start_time.minute):
        return JsonResponse({"ok": False, "error": "結束時間需晚於開始時間"}, status=400)

    start_date, end_date, _, _, _, _, _ = get_active_window()
    if shift.date < start_date or shift.date > end_date:
        return JsonResponse({"ok": False, "error": "不在可排班的日期區間內"}, status=400)

    conflict = Shift.objects.filter(
        employee=profile,
        date=shift.date,
        start_time__lt=end_time,
        end_time__gt=start_time,
        is_published=True,
    ).exclude(id=shift.id).exists()
    if conflict:
        return JsonResponse({"ok": False, "error": "排班時間重疊，請重新確認。"}, status=400)

    break_minutes = parse_break_minutes(raw_break_minutes)
    if raw_break_minutes not in (None, "") and break_minutes is None:
        return JsonResponse({"ok": False, "error": "休息時間格式錯誤"}, status=400)

    shift.start_time = start_time
    shift.end_time = end_time
    shift.store_id = None
    if break_minutes is not None:
        shift.break_minutes = break_minutes
    shift.save(update_fields=["start_time", "end_time", "store_id", "break_minutes"])

    return JsonResponse({"ok": True})


@csrf_exempt
@login_required
def delete_worker_shift(request):
    try:
        profile = request.user.userprofile
    except UserProfile.DoesNotExist:
        return JsonResponse({"ok": False, "error": "查無使用者資料"}, status=400)
    if not is_worker(request.user):
        return JsonResponse({"ok": False, "error": "僅限員工操作"}, status=403)
    if not worker_shift_edit_allowed():
        return JsonResponse({"ok": False, "error": "目前未開放員工排班"}, status=403)

    try:
        data = json.loads(request.body)
        shift_id = int(data.get("id"))
    except Exception:
        return JsonResponse({"ok": False, "error": "資料格式錯誤"}, status=400)

    shift = Shift.objects.filter(
        id=shift_id,
        employee=profile,
        is_published=True,
    ).first()
    if not shift:
        return JsonResponse({"ok": False, "error": "找不到資料"}, status=404)
    if shift.store_id:
        return JsonResponse({"ok": False, "error": "店長排定班表不可刪除"}, status=403)
    shift.delete()

    return JsonResponse({"ok": True})
